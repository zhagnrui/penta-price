"""
Build weekly_update_template.xlsx
Run: python templates/build_template.py
Generates: templates/weekly_update_template.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

# ── Styles ─────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1D9E75")   # brand green
HDR_FONT   = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
SEC_FILL   = PatternFill("solid", fgColor="E1F5EE")   # light green
SEC_FONT   = Font(name="Microsoft YaHei", size=10, bold=True, color="0F6E56")
NOTE_FONT  = Font(name="Microsoft YaHei", size=9, italic=True, color="7a9e8a")
VAL_FONT   = Font(name="Microsoft YaHei", size=10)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN       = Side(style="thin", color="D4EDE3")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_section(ws, row, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.fill = SEC_FILL
    cell.font = SEC_FONT
    cell.alignment = LEFT

def style_value_row(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = VAL_FONT
        cell.alignment = LEFT
        cell.border = BORDER

def add_note(ws, cell_ref, text):
    ws[cell_ref].comment = Comment(text, "Template")

# ── Workbook ───────────────────────────────────────────
wb = Workbook()

# ================================================================
# Sheet 0 : README (how to use)
# ================================================================
ws = wb.active
ws.title = "README 使用说明"
ws.column_dimensions["A"].width = 100

readme_lines = [
    ("季戊四醇价格周报 — 周更新模板 / Weekly Update Template", HDR_FONT, HDR_FILL),
    ("", None, None),
    ("📋 使用流程 / Workflow", SEC_FONT, SEC_FILL),
    ("1. 每周填写下面各个工作表里的价格数字和新闻。", None, None),
    ("2. 只改【黄色背景】的单元格，其他区域不要动。", None, None),
    ("3. 填完保存 .xlsx 文件，连同这个文件发给 Claude（或我）。", None, None),
    ("4. Claude 会自动转成 lib/priceData.ts，本地 build 验证后推到线上。", None, None),
    ("", None, None),
    ("📑 工作表目录 / Sheets", SEC_FONT, SEC_FILL),
    ("• Meta           — 周次标签、更新日期", None, None),
    ("• Mono 单季      — 单季戊四醇价格、地区报价、12周走势", None, None),
    ("• Di 双季        — 双季戊四醇价格、供需、18月走势", None, None),
    ("• Intl 国际      — 全球四地区价格、季度走势", None, None),
    ("• News 新闻      — 所有市场快评（标题+正文+来源链接）", None, None),
    ("", None, None),
    ("✍️ 填写原则 / Rules", SEC_FONT, SEC_FILL),
    ("• 所有价格只填【数字】，不要加 ¥、$、元、吨、逗号等符号。", None, None),
    ("  例：11850  ✔      ¥11,850  ✘      1.185万  ✘", None, None),
    ("• 单位已经在列标题里标明（¥/t 或 USD/t），别搞错币种。", None, None),
    ("• 日期格式统一用 YYYY-MM-DD，例 2026-04-17。", None, None),
    ("• 走势数组：删掉最老一个数，末尾追加本期数字 —— 保持长度不变。", None, None),
    ("• 新闻每条都要有标题+正文+日期；来源链接有就填，没有留空。", None, None),
    ("• 英文字段（textEn/bodyEn）可留空，Claude 会自动翻译。", None, None),
    ("", None, None),
    ("⚠️ 常见陷阱 / Pitfalls", SEC_FONT, SEC_FILL),
    ("• 不要用「万」做单位 —— 外国读者看不懂。11.85k 请写 11850。", None, None),
    ("• 双季戊四醇「vs 2024-10 涨幅%」必须和 baseline 对得上：", None, None),
    ("    baseline × (1 + pct/100) ≈ marketAvg", None, None),
    ("  若涨幅 175%、均价 62000，则 baseline 应填 ~22545。", None, None),
    ("• 供需单位是【万吨】—— 若数据源是 kt，先除以 10。", None, None),
    ("• history 数组长度必须和 historyLabels 一致，否则图表会断。", None, None),
    ("", None, None),
    ("💡 颜色说明 / Cell colors", SEC_FONT, SEC_FILL),
    ("  深绿 = 表头（不要改）", None, None),
    ("  浅绿 = 分节标题（不要改）", None, None),
    ("  黄色 = 需要你填写的单元格 ← 只改这些", None, None),
    ("  白色 = 说明文字（不要改）", None, None),
]
for i, (text, font, fill) in enumerate(readme_lines, start=1):
    cell = ws.cell(row=i, column=1, value=text)
    cell.alignment = LEFT
    if font: cell.font = font
    if fill: cell.fill = fill

# ================================================================
# Sheet 1 : Meta
# ================================================================
ws = wb.create_sheet("Meta")
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 50

ws["A1"] = "字段 Field"
ws["B1"] = "填写 Value"
ws["C1"] = "说明 Note"
style_header(ws, 1, [1, 2, 3])

meta_rows = [
    ("weekLabel",  "Week 16, 2026",  "本周标识，格式 'Week N, YYYY'"),
    ("updatedAt",  "2026-04-17",     "实际更新日期，格式 YYYY-MM-DD"),
]
YELLOW = PatternFill("solid", fgColor="FFF8D4")
for r, (k, v, note) in enumerate(meta_rows, start=2):
    ws.cell(row=r, column=1, value=k).font = VAL_FONT
    cell = ws.cell(row=r, column=2, value=v)
    cell.font = VAL_FONT
    cell.fill = YELLOW
    ws.cell(row=r, column=3, value=note).font = NOTE_FONT
    for c in (1, 2, 3):
        ws.cell(row=r, column=c).border = BORDER

# ================================================================
# Sheet 2 : Mono 单季
# ================================================================
ws = wb.create_sheet("Mono 单季")
for i, w in enumerate([28, 18, 40], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = 1
ws.cell(row=row, column=1, value="字段 Field")
ws.cell(row=row, column=2, value="填写 Value")
ws.cell(row=row, column=3, value="说明 Note")
style_header(ws, row, [1, 2, 3])
row += 1

# --- 基础价格 ---
style_section(ws, row, 3); ws.cell(row=row, column=1, value="■ 基础价格 / Core prices")
row += 1
mono_fields = [
    ("domesticAvg",         11850, "国内均价 ¥/吨（不含货币符号）"),
    ("weekChange",           0,    "国内均价 WoW 变化 ¥/吨，可正可负"),
    ("grade95.low",       10000,   "95% 含量区间下限 ¥/吨"),
    ("grade95.high",      10500,   "95% 含量区间上限 ¥/吨"),
    ("grade95ChangeWoW",   -200,   "95% 较上周变化 ¥/吨"),
    ("grade98.low",       13500,   "98% 含量区间下限 ¥/吨"),
    ("grade98.high",      14000,   "98% 含量区间上限 ¥/吨"),
    ("grade98ChangeWoW",   100,    "98% 较上周变化 ¥/吨"),
    ("fobQingdao",         1520,   "FOB 青岛 USD/吨"),
    ("fobChangeMoM",        -40,   "FOB 较上月变化 USD/吨"),
]
for k, v, note in mono_fields:
    ws.cell(row=row, column=1, value=k).font = VAL_FONT
    c = ws.cell(row=row, column=2, value=v)
    c.font = VAL_FONT
    c.fill = YELLOW
    ws.cell(row=row, column=3, value=note).font = NOTE_FONT
    style_value_row(ws, row, [1, 2, 3])
    row += 1

row += 1  # blank

# --- 地区报价 ---
style_section(ws, row, 3); ws.cell(row=row, column=1, value="■ 各地区报价 / Regional quotes (5 行固定)")
row += 1
ws.cell(row=row, column=1, value="name 中文名")
ws.cell(row=row, column=2, value="nameEn 英文名")
ws.cell(row=row, column=3, value="price 价格区间 / trend 趋势(up|down|flat)")
style_header(ws, row, [1, 2, 3])
row += 1

# expand columns for region section
ws.column_dimensions["D"].width = 12  # trend col

regions = [
    ("华东 95%",     "East China 95%",         "10,000–10,400", "down"),
    ("华南 95%",     "South China 95%",        "10,200–10,600", "flat"),
    ("华北 98%",     "North China 98%",        "13,500–14,000", "up"),
    ("内蒙出厂 98%", "Inner Mongolia EXW 98%", "13,200–13,800", "flat"),
    ("FOB 青岛",     "FOB Qingdao (USD/t)",    "$1,480–1,560",  "down"),
]
# Add header for 4-col region table
ws.cell(row=row-1, column=4, value="trend")
style_header(ws, row-1, [4])

# data validation for trend
dv = DataValidation(type="list", formula1='"up,down,flat"', allow_blank=False)
ws.add_data_validation(dv)

for cn, en, price, trend in regions:
    ws.cell(row=row, column=1, value=cn)
    ws.cell(row=row, column=2, value=en)
    ws.cell(row=row, column=3, value=price)
    ws.cell(row=row, column=4, value=trend)
    for c in range(1, 5):
        cell = ws.cell(row=row, column=c)
        cell.font = VAL_FONT
        cell.fill = YELLOW
        cell.border = BORDER
        cell.alignment = LEFT
    dv.add(f"D{row}")
    row += 1

row += 1

# --- 12周走势 ---
style_section(ws, row, 4); ws.cell(row=row, column=1, value="■ 12 周走势 / 12-week history (追加本周、删最老)")
row += 1
ws.cell(row=row, column=1, value="historyLabels →")
for i in range(12):
    c = ws.cell(row=row, column=2+i, value=f"W{4+i}")
    c.font = VAL_FONT; c.fill = YELLOW; c.border = BORDER; c.alignment = CENTER
row += 1
ws.cell(row=row, column=1, value="history12w →")
vals = [11200, 11400, 11100, 11600, 11800, 11500, 11900, 12100, 11850, 11700, 11900, 11850]
for i, v in enumerate(vals):
    c = ws.cell(row=row, column=2+i, value=v)
    c.font = VAL_FONT; c.fill = YELLOW; c.border = BORDER; c.alignment = CENTER
row += 1
ws.cell(row=row, column=1, value="说明").font = NOTE_FONT
ws.cell(row=row, column=2, value="删除最左 1 格，末尾填本周新值 · 两行长度必须一致").font = NOTE_FONT
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)

# ================================================================
# Sheet 3 : Di 双季
# ================================================================
ws = wb.create_sheet("Di 双季")
for i, w in enumerate([28, 18, 50], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = 1
ws.cell(row=row, column=1, value="字段 Field")
ws.cell(row=row, column=2, value="填写 Value")
ws.cell(row=row, column=3, value="说明 Note")
style_header(ws, row, [1, 2, 3])
row += 1

style_section(ws, row, 3); ws.cell(row=row, column=1, value="■ 核心指标 / Core metrics")
row += 1
di_fields = [
    ("marketAvg",            62000, "市场均价 ¥/吨"),
    ("marketAvgChangeMoM",   -3000, "市场均价 MoM 变化 ¥/吨"),
    ("highEnd",              68000, "高端上限 ¥/吨（不要写成 6.8 或 7万）"),
    ("fob",                   8400, "FOB 出口 USD/吨"),
    ("baseline2024Oct",      22500, "2024-10 基准价 ¥/吨（用来校准涨幅%）"),
    ("vsOct2024Pct",           175, "较 2024-10 涨幅 % ★ 必须和 baseline/marketAvg 对得上"),
    ("supply",                 1.8, "有效供给 万吨（注意时间口径：年/季/月）"),
    ("demand",                 2.5, "市场需求 万吨"),
]
for k, v, note in di_fields:
    ws.cell(row=row, column=1, value=k).font = VAL_FONT
    c = ws.cell(row=row, column=2, value=v)
    c.font = VAL_FONT; c.fill = YELLOW
    ws.cell(row=row, column=3, value=note).font = NOTE_FONT
    style_value_row(ws, row, [1, 2, 3])
    row += 1

# add formula check comment
add_note(ws, "B6", "校验公式：baseline2024Oct × (1 + vsOct2024Pct/100) ≈ marketAvg\n例：22500 × 2.75 = 61875 ≈ 62000 ✔")

row += 1
style_section(ws, row, 3); ws.cell(row=row, column=1, value="■ 18 个月走势 / 18-month history (追加本月、删最老)")
row += 1

# Wider layout for 18 months
ws.cell(row=row, column=1, value="historyLabels →")
labels = ['2024-10','11','12','2025-01','02','03','04','05','06','07','08','09','10','11','12','2026-01','02','03']
for i, lab in enumerate(labels):
    c = ws.cell(row=row, column=2+i, value=lab)
    c.font = VAL_FONT; c.fill = YELLOW; c.border = BORDER; c.alignment = CENTER
    ws.column_dimensions[get_column_letter(2+i)].width = 10
row += 1
ws.cell(row=row, column=1, value="history18m →")
vals18 = [25000,28000,32000,35000,42000,55000,60000,62000,58000,61000,65000,63000,60000,64000,68000,66000,62000,62000]
for i, v in enumerate(vals18):
    c = ws.cell(row=row, column=2+i, value=v)
    c.font = VAL_FONT; c.fill = YELLOW; c.border = BORDER; c.alignment = CENTER

# ================================================================
# Sheet 4 : Intl 国际
# ================================================================
ws = wb.create_sheet("Intl 国际")
for i, w in enumerate([28, 18, 50], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = 1
ws.cell(row=row, column=1, value="字段 Field")
ws.cell(row=row, column=2, value="填写 Value")
ws.cell(row=row, column=3, value="说明 Note")
style_header(ws, row, [1, 2, 3])
row += 1

style_section(ws, row, 3); ws.cell(row=row, column=1, value="■ 四地区当前价格 / Current prices (USD/t)")
row += 1
intl_fields = [
    ("us",         2090, "美国 CIF USD/吨"),
    ("europe",     2170, "欧洲 德国 USD/吨"),
    ("chinafob",   1520, "中国 FOB USD/吨"),
    ("sea",        1570, "东南亚 CIF USD/吨"),
    ("usChange",     22, "美国 MoM %"),
    ("euChange",     16, "欧洲 MoM %"),
    ("cnChange",     -6, "中国 MoM %（下跌填负数）"),
    ("seaChange",    -2, "东南亚 MoM %"),
]
for k, v, note in intl_fields:
    ws.cell(row=row, column=1, value=k).font = VAL_FONT
    c = ws.cell(row=row, column=2, value=v)
    c.font = VAL_FONT; c.fill = YELLOW
    ws.cell(row=row, column=3, value=note).font = NOTE_FONT
    style_value_row(ws, row, [1, 2, 3])
    row += 1

row += 1
style_section(ws, row, 3); ws.cell(row=row, column=1, value="■ 季度走势 / Quarterly history")
row += 1
# 11 quarterly points × 4 series
qlabels = ['Q1 23','Q2 23','Q3 23','Q4 23','Q1 24','Q2 24','Q3 24','Q4 24','Q1 25','Q3 25','Q1 26']
ws.cell(row=row, column=1, value="labels →")
for i, lab in enumerate(qlabels):
    c = ws.cell(row=row, column=2+i, value=lab)
    c.font = VAL_FONT; c.fill = YELLOW; c.border = BORDER; c.alignment = CENTER
    ws.column_dimensions[get_column_letter(2+i)].width = 10
row += 1

series = [
    ("us",  [1800,1750,2160,1900,1800,1850,2160,1900,2090,2090,2090]),
    ("eu",  [1700,1650,1945,1800,1700,1700,1945,1820,2170,2000,2170]),
    ("cn",  [1500,1450,1447,1400,1380,1530,1447,1380,1620,1500,1520]),
    ("sea", [1450,1380,1400,1350,1320,1470,1400,1370,1570,1520,1570]),
]
for name, vals in series:
    ws.cell(row=row, column=1, value=f"{name} →")
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=2+i, value=v)
        c.font = VAL_FONT; c.fill = YELLOW; c.border = BORDER; c.alignment = CENTER
    row += 1

# ================================================================
# Sheet 5 : News 新闻
# ================================================================
ws = wb.create_sheet("News 新闻")
headers = [
    ("panel",      10, "mono / di / intl"),
    ("tag",        12, "中文标签，如 供给 / 成本 / PCB需求"),
    ("tagEn",      14, "英文标签（可空）"),
    ("text",       40, "中文一句话标题（≤60字）"),
    ("textEn",     40, "英文一句话标题（可空）"),
    ("body",       60, "中文正文 2–4 句。可用 Alt+Enter 换行"),
    ("bodyEn",     60, "英文正文（可空，Claude 可自动译）"),
    ("date",       12, "YYYY-MM-DD"),
    ("source",     18, "来源机构名，如 卓创资讯 / ICIS"),
    ("url",        40, "来源链接（可空；填了读者可点击跳转）"),
]
for i, (name, width, _) in enumerate(headers, start=1):
    ws.cell(row=1, column=i, value=name)
    ws.column_dimensions[get_column_letter(i)].width = width
style_header(ws, 1, list(range(1, len(headers)+1)))

# add notes to header row
for i, (_, _, note) in enumerate(headers, start=1):
    add_note(ws, f"{get_column_letter(i)}1", note)

# dv for panel column
dv_panel = DataValidation(type="list", formula1='"mono,di,intl"', allow_blank=False)
ws.add_data_validation(dv_panel)

# sample rows
samples = [
    ("mono", "供给", "Supply",
     "湖北宜化宜昌装置升级改造进入关键阶段，供给端关注度持续升温",
     "Hubei Yihua Yichang plant upgrade enters a critical phase; supply-side attention is building.",
     "宜化宜昌基地 3 万吨/年单季戊四醇装置自 3 月下旬进入升级改造，预计 5 月中旬复产。该装置占华中地区 95% 含量总产能约 22%，短期内华中现货供给边际收紧。下游贸易商反映提货周期已从 3 天延长至 7 天。",
     "The 30 kt/y mono-PE line at Yihua's Yichang base entered a scheduled upgrade in late March...",
     "2026-04-10", "", ""),
    ("mono", "成本", "Cost",
     "甲醛原料价格本周小幅回落，对单季成本端形成短期压力",
     "Formaldehyde feedstock eased slightly this week...",
     "本周华东 37% 甲醛主流报价下调 40–60 元/吨至 1,280–1,340 元/吨...",
     "",
     "2026-04-10", "", ""),
    ("di", "PCB需求", "PCB demand",
     "双季供需缺口持续，PCB光固化油墨需求驱动高端报价持续支撑",
     "",
     "高端 90% 含量双季戊四醇主要用于 PCB 光固化油墨和高端醇酸树脂...",
     "",
     "2026-04-10", "", ""),
    ("intl", "北美", "North America",
     "美国市场3月环比上涨22%，受Perstorp检修收紧供应 + 关税前抢货推动",
     "US market surged 22% MoM in March 2025...",
     "Perstorp 瑞典工厂 3 月下旬开始为期 4 周的计划性检修...",
     "",
     "2026-04-10", "", ""),
]
for r, row_data in enumerate(samples, start=2):
    for c, val in enumerate(row_data, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = VAL_FONT
        cell.fill = YELLOW
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    dv_panel.add(f"A{r}")
    ws.row_dimensions[r].height = 60

# add empty rows for user to fill
for r in range(len(samples)+2, len(samples)+10):
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = YELLOW
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    dv_panel.add(f"A{r}")
    ws.row_dimensions[r].height = 40

# Freeze panes on data sheets
for sheet_name in ["Mono 单季", "Di 双季", "Intl 国际", "News 新闻"]:
    wb[sheet_name].freeze_panes = "A2"

# Save
out = "D:/Training/Claude/penta-price/templates/weekly_update_template.xlsx"
wb.save(out)
print(f"Saved: {out}")
